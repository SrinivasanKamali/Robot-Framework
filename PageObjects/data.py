from openpyxl import load_workbook

class data:
    def read_excel_cell(self,file_path,sheet_name,row,column):
        book=load_workbook(file_path)
        sheet=book[sheet_name]
        headers=[cell.value for cell in sheet[1]]
        col_index=headers.index(column)+1
        return sheet.cell(row=int(row),column=col_index).value
    def get_max_row(self,file_path,sheet_name):
        book = load_workbook(file_path)
        sheet = book[sheet_name]
        return sheet.max_row